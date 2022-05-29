from copy import copy
import librosa
import argparse
import numpy as np
import openpyxl
from openpyxl import Workbook
import tensorflow as tf
import time
import json
from os import listdir
from pathlib import Path
from datetime import datetime
from tensorflow.keras import models
from tensorflow.keras.layers import Input
from audio_preprocessing import preprocess_audio, METHODS
from tensorflow.keras.applications.resnet import ResNet50
from tensorflow.keras.applications.densenet import DenseNet121
from tensorflow.keras.applications.inception_v3 import InceptionV3
from distance_visualization import display_results

MODELS = ['ResNet', 'DenseNet', 'Inception']
SHAPE = (128, 128, 3)


class FeatureExtractionModel:
    """
    a sample class to utilize pretrained fine-tuned tf models
    """

    def __init__(self, name=MODELS[0], weights='imagenet'):
        """
        initializes a model
        :param name: name of the model
        :param weights: a path to its weights
        """
        if name == MODELS[0]:
            base_model = ResNet50(weights=weights, input_tensor=Input(shape=SHAPE), include_top=False)
        elif name == MODELS[1]:
            base_model = DenseNet121(weights=weights, input_tensor=Input(shape=SHAPE), include_top=False)
        elif name == MODELS[2]:
            base_model = InceptionV3(weights=weights, input_tensor=Input(shape=SHAPE), include_top=False)
        else:
            raise ValueError(
                f"{datetime.now()}: Unsupported model. Supported models are {MODELS}'.")
        global_average_layer = tf.keras.layers.GlobalAveragePooling2D()
        final_model = models.Sequential()
        final_model.add(base_model)
        final_model.add(global_average_layer)
        self.model = final_model

    @staticmethod
    def __validate_data(data):
        if len(data.shape) == 4 and data.shape[1] == 128 and data.shape[2] == 128 and data.shape[3] == 3:
            return True
        else:
            return False

    def get_feature_vectors(self, data):
        if self.__validate_data(data):
            return self.model.predict(data)


def load_audio(path):
    return librosa.load(path)


def process_audio(audio, rate, preproc, model=MODELS[0], weights='imagenet'):
    fem = FeatureExtractionModel(model, weights)
    preprocessed = preprocess_audio(audio, rate, preproc)
    spectrogram = np.pad(preprocessed, ((0, 0), (0, 128 - (preprocessed.shape[1]) % 128), (0, 0)), 'empty')
    segments = [spectrogram[:, y:y + 128, :] for y in range(0, spectrogram.shape[1], 64)]
    result = fem.get_feature_vectors(np.stack(segments[0: 20]))
    print(f'{datetime.now()}: got {result.shape[0]} {result[0].shape[0]}-dimensional vectors from {model}')
    return result

def norm_calc(audios, trek):
    dt = []
    for fl in audios:
        dt.append([load_audio(fl)]) 
    for net_mod in MODELS:
        book_tmp = openpyxl.Workbook()
        mean_met1_1 = []
        mean_met1_2 = []
        mean_met2_1 = []
        mean_met2_2 = [] 
        for index_of_audios in range(len(dt) - 1):
            sheet_1 = book.create_sheet(f"{net_mod}_results", 0)
            arr_1st_method_1 = process_audio(dt[index_of_audios], dt[index_of_audios+1], 'guzhov', net_model)
            arr_2nd_method_1 = process_audio(dt[index_of_audios], dt[index_of_audios+1], 'palanisamy', net_model)
            arr_1st_method_2 = process_audio(dt[index_of_audios], dt[index_of_audios+1], 'guzhov', net_model)
            arr_2nd_method_2 = process_audio(dt[index_of_audios], dt[index_of_audios+1], 'palanisamy', net_model)
            met1_1 = np.array([])
            met1_2 = np.array([])
            met2_1 = np.array([])
            met2_2 = np.array([])
            for n in range(11):
                met1_1 = np.append(met1_1, np.linalg.norm(arr_1st_method_1[n] - arr_1st_method_1[n + 1]))
                met1_1 = np.append(met1_1, np.linalg.norm(arr_1st_method_2[n] - arr_1st_method_2[n + 1]))
                met1_2 = np.append(met1_2, np.linalg.norm(arr_1st_method_1[n] - arr_1st_method_2[n]))
                met2_2 = np.append(met2_2, np.linalg.norm(arr_2nd_method_1[n] - arr_2nd_method_2[n]))
                met2_1 = np.append(met2_1, np.linalg.norm(arr_2nd_method_1[n] - arr_2nd_method_1[n + 1]))
                met2_1 = np.append(met2_1, np.linalg.norm(arr_2nd_method_2[n] - arr_2nd_method_2[n + 1]))
            mean_met1_1.append([np.mean(met1_1.tolist())])
            mean_met1_2.append([np.mean(met1_2.tolist())])
            mean_met2_1.append([np.mean(met2_1.tolist())])
            mean_met2_2.append([np.mean(met2_2.tolist())])
        sheet_1.append(['guzhov, same'])
        sheet_1.append([np.mean(mean_met1_1.tolist())])
        sheet_1.append(['guzhov, different'])
        sheet_1.append([np.mean(mean_met1_2.tolist())])
        sheet_1.append(['palanisamy, same'])
        sheet_1.append([np.mean(mean_met2_1.tolist())])
        sheet_1.append(['palanisamy, different'])
        sheet_1.append([np.mean(mean_met2_2.tolist())])
        book_tmp.save(f'{trek}'+f'{net_mod}_results.xlsx')

def time_for_diff_models(audios, trek):
    dt = []
    t1 = []
    t2 = []
    for fl in audios:
        dt.append([load_audio(fl)])
    book = openpyxl.Workbook()
    sheet_1 = book.create_sheet("Results_of_testing_different_models", 0)
    sheet_1.append(['Метод предобработки', 'Модель нейросети', 'Среднее время обработки одного фрагмента фонограммы'])
    for index_of_audios in range(len(dt) - 1):
        for net_model in MODELS:
            startTime = time.time()
            process_audio(dt[index_of_audios], dt[index_of_audios+1], 'guzhov', net_model)
            endTime = time.time()
            totalTime = endTime - startTime
            print('Время выполнения для модели ' + net_model + ' методом "Guzhov et al." равно: ' + str(totalTime))
            t1.append(totalTime)
            startTime = time.time()
            process_audio(dt[index_of_audios], dt[index_of_audios+1], 'palanisamy', net_model)
            endTime = time.time()
            totalTime = endTime - startTime
            print('Время выполнения для модели ' + net_model + ' методом "Palanisamy et al." равно: ' + str(totalTime))
            t2.append(totalTime)
    t1.append(['Guzhov et al.', net_model, totalTime])
    t2.append(['Palanisamy et al.', net_model, totalTime])
    book.save(f'{trek}'+'\Results_of_testing_different_models.xlsx')



if __name__ == "__main__":
    SETTINGS = 'C:\\Users\\stron\\Documents\\GitHub\\docs\\params\\main_args.json'
    # parser = argparse.ArgumentParser(description="feature vector extracting script")
    # parser.add_argument("-r", dest="research", action='store_true', help="runs preset research if provided")
    # parser.add_argument("-m", dest="model", type=str, choices=MODELS, help="name of the model to use")
    # parser.add_argument("-w", dest="weights", type=str, default='imagenet', help="path to the model weights")
    # parser.add_argument("-a", dest="audio", type=str, help="path to an audio file")
    # parser.add_argument("-p", dest="preproc", type=str, choices=METHODS, default='palanisamy', help="name of the "
    #                                                                                                 "preprocessing "
    #                                                                                                 "method")
    # args = parser.parse_args()

    #these two lines are needed to run inference on my dated gtx 1660 ti
    #physical_devices = tf.config.experimental.list_physical_devices('GPU')
    #tf.config.experimental.set_memory_growth(physical_devices[0], True)
    with open(SETTINGS) as json_file:
        settings = json.load(json_file)
    time_for_diff_models(settings["audios"], settings["trek"])
    norm_calc(settings["audios"], settings["trek"])